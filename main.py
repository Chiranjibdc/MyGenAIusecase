import io
import re
import json
from datetime import date
from urllib.parse import urlparse

import streamlit as st
import pandas as pd

# ── page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Gen AI Use Case & Fitment Finder",
    page_icon="🤖",
    layout="wide",
)

# ── helpers ───────────────────────────────────────────────────────────────────

def normalize_url(raw: str) -> str:
    raw = raw.strip()
    if not raw.startswith(("http://", "https://")):
        raw = "https://" + raw
    return raw


def extract_domain(url: str) -> str:
    parsed = urlparse(url)
    domain = parsed.netloc or parsed.path
    domain = re.sub(r"^www\.", "", domain)
    return domain.split("/")[0]


# ── Claude research ───────────────────────────────────────────────────────────

def run_research(url: str, api_key: str, extra_context: str = "") -> dict:
    from anthropic import Anthropic

    client = Anthropic(api_key=api_key)
    domain = extract_domain(url)

    web_search_tool = {
        "type": "web_search_20250305",
        "name": "web_search",
        "max_uses": 10,
    }

    submit_tool = {
        "name": "submit_use_cases",
        "description": (
            "Call this tool exactly once after completing all research to submit the "
            "final structured Gen AI use case report. Do NOT call it before you have "
            "finished all web searches."
        ),
        "input_schema": {
            "type": "object",
            "required": [
                "company_name", "company_summary", "it_landscape_summary",
                "fitment_assessment", "fitment_score", "use_cases",
            ],
            "properties": {
                "company_name": {"type": "string"},
                "company_summary": {"type": "string"},
                "it_landscape_summary": {"type": "string"},
                "fitment_assessment": {"type": "string"},
                "fitment_score": {"type": "integer", "minimum": 1, "maximum": 10},
                "use_cases": {
                    "type": "array",
                    "minItems": 20,
                    "maxItems": 20,
                    "items": {
                        "type": "object",
                        "required": [
                            "id", "use_case", "scope", "function_area", "category",
                            "complexity", "cost_build_usd", "cost_run_usd_per_year",
                            "timeline", "description", "fitment_rationale",
                        ],
                        "properties": {
                            "id": {"type": "integer"},
                            "use_case": {"type": "string"},
                            "scope": {
                                "type": "string",
                                "enum": ["Internal", "Client-facing", "Internal + Client-facing"],
                            },
                            "function_area": {"type": "string"},
                            "category": {
                                "type": "string",
                                "enum": ["Quick Win", "Medium Lift", "Strategic Bet"],
                            },
                            "complexity": {
                                "type": "string",
                                "enum": ["Low", "Low-Medium", "Medium", "Medium-High", "High"],
                            },
                            "cost_build_usd": {"type": "string"},
                            "cost_run_usd_per_year": {"type": "string"},
                            "timeline": {"type": "string"},
                            "description": {"type": "string"},
                            "fitment_rationale": {"type": "string"},
                        },
                    },
                },
            },
        },
    }

    system_prompt = f"""You are a senior Gen AI strategy and IT research analyst.
Your task is to deeply research a target company using web search, then identify the top 20 highest-impact Generative AI use cases tailored specifically to that company.

TARGET COMPANY URL: {url}
TARGET DOMAIN: {domain}
{"ADDITIONAL CONTEXT FROM USER: " + extra_context if extra_context.strip() else ""}

RESEARCH INSTRUCTIONS:
1. Start by searching the company's OWN website first (use queries like "site:{domain} about" and "site:{domain} investor newsroom technology").
2. Then broaden to general web searches covering:
   - Recent news, press releases, and leadership statements about the company
   - Technology/IT-specific coverage (ERP, CRM, cloud, data platforms in use)
   - Job postings that mention AI, ML, data engineering, cloud roles (strong signal of active initiatives)
   - Earnings call transcripts or annual reports mentioning digital transformation, AI pilots, or technology strategy
   - Analyst commentary on the company's digital maturity

3. Aim for 6-10 search calls total, covering at minimum:
   a) Company about/overview page
   b) Investor relations / annual report / newsroom
   c) Technology stack signals (news + job postings)
   d) AI/Gen AI strategy or pilot announcements
   e) Pain points, business challenges, recent strategic shifts

4. Research the following dimensions:
   - Core business model, industry, and primary revenue drivers
   - Current IT landscape and known technology stack (ERP, CRM, cloud provider, data platforms)
   - Digital transformation initiatives and stated AI/Gen AI strategy or pilots
   - Organizational scale and structure (customer service volume, manufacturing ops, financial ops, etc.)
   - Budget signals, regulatory environment, and leadership tone toward technology

5. Assess "Gen AI Fitment": based on your research, judge how ready and receptive this company is to Gen AI adoption (digital maturity, prior AI investments, regulatory constraints, budget signals, leadership tone). Score 1-10.

6. Identify the TOP 20 most business-impactful Gen AI use cases for THIS specific company — not generic industry boilerplate. Each use case must:
   - Be grounded in at least one specific research signal you found
   - Span a mix of internal operations and customer/client-facing opportunities
   - Be weighted toward the company's evidenced IT maturity and strategic direction
   - Include realistic cost and timeline estimates given the company's likely scale

7. When you have finished ALL web searches and are ready to submit, call the `submit_use_cases` tool exactly once with the complete structured result. Do NOT call it before you are done researching.
"""

    messages = [
        {
            "role": "user",
            "content": (
                f"Please research {url} and identify the top 20 Gen AI use cases for this company. "
                "Use web search extensively to understand their IT landscape, technology roadmap, and strategic direction. "
                "When done, call the submit_use_cases tool with your complete structured findings."
            ),
        }
    ]

    tools = [web_search_tool, submit_tool]

    status_placeholder = st.empty()
    result_data = None
    raw_fallback = []

    with st.status("Researching company...", expanded=True) as status_box:
        for turn in range(15):
            response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=8000,
                system=system_prompt,
                tools=tools,
                messages=messages,
            )

            # collect assistant content
            assistant_content = response.content
            messages.append({"role": "assistant", "content": assistant_content})

            tool_results = []
            found_submit = False

            for block in assistant_content:
                btype = getattr(block, "type", None)

                if btype == "text":
                    raw_fallback.append(block.text)

                elif btype == "tool_use":
                    if block.name == "submit_use_cases":
                        status_box.write("✅ Compiling final use case report...")
                        result_data = block.input
                        found_submit = True
                        # provide a tool result so the conversation is valid
                        tool_results.append({
                            "type": "tool_result",
                            "tool_use_id": block.id,
                            "content": "Report submitted successfully.",
                        })

                    elif block.name == "web_search":
                        # web_search is server-side; input has a "query" key
                        query = getattr(block, "input", {}).get("query", "")
                        status_box.write(f"🔍 Searching: {query}")
                        # No client-side execution needed — result will be in next response

                elif btype == "tool_result":
                    pass  # server-side results embedded by the API

            if found_submit:
                break

            # if there are pending client tool results, add them
            if tool_results:
                messages.append({"role": "user", "content": tool_results})
            elif response.stop_reason in ("end_turn", "stop_sequence"):
                # Claude stopped without calling submit_use_cases
                break
            # otherwise loop continues (e.g. stop_reason == "tool_use" for web_search)

        if result_data:
            status_box.update(label="Research complete!", state="complete")
        else:
            status_box.update(label="Research ended without structured output.", state="error")

    if not result_data:
        raise ValueError(
            "Claude did not call submit_use_cases within the turn limit. "
            "Raw output:\n\n" + "\n".join(raw_fallback)
        )

    use_cases = result_data.get("use_cases", [])
    if len(use_cases) != 20:
        raise ValueError(
            f"Expected exactly 20 use cases but received {len(use_cases)}. "
            "Raw output:\n\n" + json.dumps(result_data, indent=2)
        )

    return result_data


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_excel(data: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    wb = openpyxl.Workbook()

    # ── Sheet 1: Company Profile ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Company Profile"
    profile_rows = [
        ("Company Name", data["company_name"]),
        ("Company Summary", data["company_summary"]),
        ("IT Landscape Summary", data["it_landscape_summary"]),
        ("Fitment Assessment", data["fitment_assessment"]),
        ("Fitment Score (1-10)", data["fitment_score"]),
    ]
    for row_idx, (label, value) in enumerate(profile_rows, start=1):
        ws1.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        cell = ws1.cell(row=row_idx, column=2, value=str(value))
        cell.alignment = Alignment(wrap_text=True)
    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 90

    # ── Sheet 2: Top 20 Use Cases ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Top 20 Use Cases")
    headers = [
        "id", "use_case", "scope", "function_area", "category",
        "complexity", "cost_build_usd", "cost_run_usd_per_year",
        "timeline", "description", "fitment_rationale",
    ]
    col_widths = [5, 32, 18, 24, 13, 13, 14, 16, 11, 50, 40]

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws2.cell(row=1, column=col_idx, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws2.freeze_panes = "A2"

    wrap_cols = {headers.index("description") + 1, headers.index("fitment_rationale") + 1}
    for row_idx, uc in enumerate(data["use_cases"], start=2):
        for col_idx, hdr in enumerate(headers, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=uc.get(hdr, ""))
            if col_idx in wrap_cols:
                cell.alignment = Alignment(wrap_text=True)

    # ── Sheet 3: Summary ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "Category"
    ws3["B1"] = "Count"
    ws3["A1"].font = Font(bold=True)
    ws3["B1"].font = Font(bold=True)
    categories = ["Quick Win", "Medium Lift", "Strategic Bet"]
    for i, cat in enumerate(categories, start=2):
        ws3.cell(row=i, column=1, value=cat)
        ws3.cell(
            row=i, column=2,
            value=f'=COUNTIF(\'Top 20 Use Cases\'!E:E,A{i})',
        )
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 10

    wb.save(buf)
    return buf.getvalue()


# ── PDF builder ───────────────────────────────────────────────────────────────

def build_pdf(data: dict, url: str) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, HRFlowable, Table, TableStyle,
    )
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "Title2", parent=styles["Title"], fontSize=18, spaceAfter=6, alignment=TA_CENTER
    )
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=13, spaceBefore=12, spaceAfter=4)
    h2 = ParagraphStyle(
        "H2", parent=styles["Heading2"], fontSize=11, spaceBefore=10, spaceAfter=3,
        textColor=colors.HexColor("#1F3864"),
    )
    body = ParagraphStyle("Body2", parent=styles["Normal"], fontSize=9, spaceAfter=3, leading=13)
    italic = ParagraphStyle(
        "Italic2", parent=styles["Normal"], fontSize=8.5, fontName="Helvetica-Oblique",
        textColor=colors.HexColor("#555555"), spaceAfter=4, leading=12,
    )
    meta_style = ParagraphStyle(
        "Meta", parent=styles["Normal"], fontSize=8, textColor=colors.grey, alignment=TA_CENTER
    )

    category_colors = {
        "Quick Win": colors.HexColor("#D6EAD6"),
        "Medium Lift": colors.HexColor("#FFF3CD"),
        "Strategic Bet": colors.HexColor("#FADBD8"),
    }

    story = []

    # ── Page 1: header + company profile ─────────────────────────────────────
    story.append(Paragraph("Gen AI Use Case &amp; Fitment Report", title_style))
    story.append(Paragraph(
        f"<b>Company:</b> {data['company_name']} &nbsp;|&nbsp; "
        f"<b>URL:</b> {url} &nbsp;|&nbsp; "
        f"<b>Generated:</b> {date.today().isoformat()}",
        meta_style,
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1F3864"), spaceAfter=8))

    story.append(Paragraph("Company Overview", h1))
    story.append(Paragraph(data["company_summary"], body))

    story.append(Paragraph("IT Landscape &amp; Technology Signals", h1))
    story.append(Paragraph(data["it_landscape_summary"], body))

    story.append(Paragraph("Gen AI Fitment Assessment", h1))
    story.append(Paragraph(
        f"<b>Fitment Score: {data['fitment_score']}/10</b>", body
    ))
    story.append(Paragraph(data["fitment_assessment"], body))
    story.append(Spacer(1, 0.4 * cm))

    # ── Use cases grouped by category ────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1F3864"), spaceAfter=6))
    story.append(Paragraph("Top 20 Gen AI Use Cases", title_style))

    grouped: dict[str, list] = {"Quick Win": [], "Medium Lift": [], "Strategic Bet": []}
    for uc in data["use_cases"]:
        grouped.setdefault(uc["category"], []).append(uc)

    for cat in ["Quick Win", "Medium Lift", "Strategic Bet"]:
        items = grouped.get(cat, [])
        if not items:
            continue
        story.append(Paragraph(f"● {cat}", h2))
        for uc in items:
            story.append(Paragraph(
                f"<b>{uc['id']}. {uc['use_case']}</b>", body
            ))
            story.append(Paragraph(
                f"Scope: {uc['scope']} &nbsp;|&nbsp; "
                f"Function: {uc['function_area']} &nbsp;|&nbsp; "
                f"Complexity: {uc['complexity']} &nbsp;|&nbsp; "
                f"Build: {uc['cost_build_usd']} &nbsp;|&nbsp; "
                f"Run/yr: {uc['cost_run_usd_per_year']} &nbsp;|&nbsp; "
                f"Timeline: {uc['timeline']}",
                italic,
            ))
            story.append(Paragraph(uc["description"], body))
            story.append(Paragraph(f"<i>Signal: {uc['fitment_rationale']}</i>", italic))
            story.append(Spacer(1, 0.2 * cm))

    doc.build(story)
    return buf.getvalue()


# ── Streamlit UI ──────────────────────────────────────────────────────────────

def main():
    st.title("🤖 Gen AI Use Case & Fitment Finder")
    st.caption(
        "Enter any company's website URL. The app researches their IT landscape and "
        "technology roadmap using Claude's live web search, scores Gen AI readiness, "
        "and returns the top 20 tailored use cases — categorised, costed, and downloadable."
    )

    # ── Sidebar ───────────────────────────────────────────────────────────────
    with st.sidebar:
        st.header("Configuration")
        default_key = ""
        try:
            default_key = st.secrets.get("ANTHROPIC_API_KEY", "")
        except Exception:
            pass
        api_key = st.text_input(
            "Anthropic API Key",
            value=default_key,
            type="password",
            help="Not stored anywhere — used only for this session.",
        )
        st.caption("ℹ️ Your key is never stored server-side.")

        extra_context = st.text_area(
            "Additional context (optional)",
            placeholder='e.g. "focus on BFSI regulatory angles" or "company recently had a data breach"',
            height=100,
        )

        st.divider()
        st.caption("**Cost note:** each run uses Claude Sonnet with web search. "
                   "Web search is billed at $10/1K searches plus token costs. "
                   "A typical run uses ~6-10 searches.")

    # ── Main input ────────────────────────────────────────────────────────────
    url_input = st.text_input(
        "Company website URL",
        placeholder="https://www.example.com",
    )
    run_btn = st.button("🔍 Research & Generate Use Cases", type="primary")

    if run_btn:
        # Validation
        if not url_input.strip():
            st.warning("Please enter a company website URL.")
            st.stop()

        if not api_key:
            st.error(
                "Anthropic API key is required. Enter it in the sidebar, or set "
                "`ANTHROPIC_API_KEY` in `.streamlit/secrets.toml`."
            )
            st.stop()

        url = normalize_url(url_input)
        domain = extract_domain(url)

        try:
            data = run_research(url, api_key, extra_context)
        except ValueError as ve:
            msg = str(ve)
            # check if there's a raw payload embedded after a newline block
            parts = msg.split("\n\n", 1)
            st.error(f"Research did not produce a valid structured result: {parts[0]}")
            if len(parts) > 1:
                with st.expander("Raw output from Claude"):
                    st.text(parts[1])
            st.stop()
        except Exception as exc:
            st.error(f"API error: {exc}")
            st.stop()

        # ── Company profile ───────────────────────────────────────────────────
        st.subheader(f"📋 {data['company_name']}")

        with st.container(border=True):
            st.info(f"**Business Overview**\n\n{data['company_summary']}")
            st.info(f"**IT Landscape & Technology Signals**\n\n{data['it_landscape_summary']}")

        # ── Fitment score ─────────────────────────────────────────────────────
        score = int(data["fitment_score"])
        col_score, col_assess = st.columns([1, 3])
        with col_score:
            st.metric("Gen AI Fitment Score", f"{score} / 10")
            st.progress(score / 10)
        with col_assess:
            st.markdown(f"**Fitment Assessment**\n\n{data['fitment_assessment']}")

        st.divider()

        # ── Category metrics ──────────────────────────────────────────────────
        uc_list = data["use_cases"]
        qw = sum(1 for u in uc_list if u["category"] == "Quick Win")
        ml = sum(1 for u in uc_list if u["category"] == "Medium Lift")
        sb = sum(1 for u in uc_list if u["category"] == "Strategic Bet")

        m1, m2, m3 = st.columns(3)
        m1.metric("⚡ Quick Win", qw)
        m2.metric("🔧 Medium Lift", ml)
        m3.metric("🚀 Strategic Bet", sb)

        # ── Use-case table ────────────────────────────────────────────────────
        df = pd.DataFrame(uc_list)
        # Merge fitment_rationale into description for cleaner display
        df["description_full"] = df["description"] + "\n\n📌 " + df["fitment_rationale"]
        display_cols = [
            "id", "use_case", "scope", "function_area", "category",
            "complexity", "cost_build_usd", "cost_run_usd_per_year",
            "timeline", "description_full",
        ]
        df_display = df[display_cols].rename(columns={
            "id": "#",
            "use_case": "Use Case",
            "scope": "Scope",
            "function_area": "Function Area",
            "category": "Category",
            "complexity": "Complexity",
            "cost_build_usd": "Build Cost",
            "cost_run_usd_per_year": "Run Cost/yr",
            "timeline": "Timeline",
            "description_full": "Description & Signal",
        })

        cat_colors = {
            "Quick Win": "background-color: #D6EAD6",
            "Medium Lift": "background-color: #FFF3CD",
            "Strategic Bet": "background-color: #FADBD8",
        }

        def color_category(val):
            return cat_colors.get(val, "")

        styled = df_display.style.map(color_category, subset=["Category"])

        st.dataframe(
            styled,
            use_container_width=True,
            column_config={
                "#": st.column_config.NumberColumn(width="small"),
                "Use Case": st.column_config.TextColumn(width="medium"),
                "Description & Signal": st.column_config.TextColumn(width="large"),
            },
            height=700,
        )

        # ── Downloads ─────────────────────────────────────────────────────────
        st.divider()
        st.subheader("📥 Download Report")
        dl1, dl2 = st.columns(2)

        xlsx_bytes = build_excel(data)
        with dl1:
            st.download_button(
                label="⬇️ Download Excel (.xlsx)",
                data=xlsx_bytes,
                file_name=f"genai_fitment_{domain}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        pdf_bytes = build_pdf(data, url)
        with dl2:
            st.download_button(
                label="⬇️ Download PDF (.pdf)",
                data=pdf_bytes,
                file_name=f"genai_fitment_{domain}.pdf",
                mime="application/pdf",
            )

    st.divider()
    st.caption(
        "Powered by Claude (Anthropic) with live web search. No data is stored server-side."
    )


if __name__ == "__main__":
    main()
